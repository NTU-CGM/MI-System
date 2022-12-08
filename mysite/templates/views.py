from django.shortcuts import render
from mysite import models, forms
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from wsgiref.util import FileWrapper
from django.contrib.sessions.models import Session
from django.contrib import messages, auth
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth.models import User
from django.http import FileResponse
import threading
from asgiref.sync import async_to_sync
import time
import datetime
import subprocess
import os
import openpyxl
import pandas as pd
import codecs
import numpy as np
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter


#功能----------------------------------------------
def txt_to_xlsx(filename,outfile):
 
    fr = codecs.open(filename,'r')
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws = wb.create_sheet()
    ws.title = 'Sheet1'
    row = 0
    for line in fr:
        row +=1
        line = line.strip()
        line = line.split(' ')
        col = 0
        for j in range(len(line)):
            col +=1
            #print (line[j])
            ws.cell(column = col,row = row,value = line[j].format(get_column_letter(col)))
    wb.save(outfile)

def read_xlsx(filename):
    
    wb = openpyxl.load_workbook(filename)

    ws = wb.get_sheet_by_name('Sheet1')

    for row in ws.rows:
        for cell in row:
            print (cell.value)

    for col in ws.columns:
        for cell in col:
            print (cell.value)

def is_number(s):
    try:  
        float(s)
        if s%1 != 0:
            return False
        else:
            return True
    except ValueError:  # ValueError 傳入無效參數
        pass  
    try:
        import unicodedata 
        for i in s:
            unicodedata.numeric(i)  # 
            #return True
        return True
    except (TypeError, ValueError):
        pass
    return False

# Create your views here.
def index(request, pid=None, del_pass=None):
    if request.user.is_authenticated:
        username = request.user.username
    messages.get_messages(request)
    return render(request, 'index.html', locals())

def register(request, pid=None, del_pass=None):
    if request.method =='POST':
        register_form = forms.Registerform(request.POST)
        if register_form.is_valid():
            new_user = register_form.save()
            new_user.set_password(new_user.password)
            new_user.save()
           # authenticated_user = authenticate(username=new_user.username,password=request.POST['password2'])
            auth.login(request, new_user)
            messages.add_message(request, messages.SUCCESS,'註冊成功')
            return HttpResponseRedirect('/')
        else:
            messages.add_message(request, messages.INFO, '請檢查輸入的欄位內容')
    else:
        register_form = forms.Registerform()
    return render(request, 'register.html', locals())


def login(request):
    if request.method =='POST':
        login_form = forms.Loginform(request.POST)
        if login_form.is_valid():
            login_name = request.POST['username'].strip()
            login_password = request.POST['password']
            user = authenticate(username=login_name, password=login_password)
            if user is not None:
                if user.is_active:
                    auth.login(request, user)
                    messages.add_message(request, messages.SUCCESS,'登入成功')
                    return HttpResponseRedirect('/')
                else:
                    messages.add_message(request, messages.WARNING,'帳號尚未啟用')
            else:
                messages.add_message(request, messages.WARNING,'無此帳號')
        else:
            messages.add_message(request, messages.INFO, '請檢查輸入的欄位內容')
    else:
        login_form = forms.Loginform()

    return render(request, 'login.html', locals())

def logout(request):
    auth.logout(request)
    messages.add_message(request, messages.INFO, "登出成功")
    return HttpResponseRedirect('/')

#修改-------------------------------------------------------
@login_required(login_url='/login')
def imputation_page(request):
    if request.user.is_authenticated:
        username = request.user.username

    if request.POST:
        form = forms.Optiontest(request.POST, request.FILES)
        if form.is_valid():
             message = "已收到資料"
             #表單資料
             user_dir = "../../../work1782/evan/sample_dir/"+username
             if not os.path.isdir(user_dir):
                 os.mkdir(user_dir)

             project_name = form.cleaned_data['user_name']
             path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
             num = 2
             while os.path.isdir(path2):
                 project_name = project_name+"_"+str(num)
                 path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
                 num+=1
                 if num > 50 :
                     print('error')
                     break
             
             
             handle_uploaded_bedfile(request.FILES['bedfile'],username,project_name)
             handle_uploaded_bimfile(request.FILES['bimfile'],username,project_name)
             handle_uploaded_famfile(request.FILES['famfile'],username,project_name)
             user_chromosome = form.cleaned_data['user_chromosome']
             user_entirechr = form.cleaned_data['user_entirechr']
             user_position1 = form.cleaned_data['user_position1']
             user_position2 = form.cleaned_data['user_position2']
             user_mafoption = form.cleaned_data['user_mafoption']
             user_genooption = form.cleaned_data['user_genooption']
             user_mindoption = form.cleaned_data['user_mindoption']
             user_hwe = form.cleaned_data['user_hwe']
             user_group = form.cleaned_data['user_group']
             if user_entirechr == False:
                 if user_position1 == '' or user_position2 =='':
                     message = 'Please check the column of chromosome position'
                     messages.add_message(request, messages.WARNING, message)
                     return HttpResponseRedirect('/imputation/')
                 elif is_number(user_position1) == False or is_number(user_position2) ==False:
                     message = 'Please check the column of chromosome position'
                     messages.add_message(request, messages.WARNING, message)
                     return HttpResponseRedirect('/imputation/')
            
              
#session 測試------------------------------------------------------
             request.session['project_name'] = project_name
             request.session['user_chromosome'] = user_chromosome
             request.session['user_entirechr'] = user_entirechr
             request.session['user_position1'] = user_position1
             request.session['user_position2'] = user_position2
             request.session['user_mafoption'] = user_mafoption
             request.session['user_genooption'] = user_genooption
             request.session['user_mindoption'] = user_mindoption
             request.session['user_hwe'] = user_hwe
             request.session['user_group'] = user_group
             
             return HttpResponseRedirect('/submit_page/')
             

            
        else:
             message = "請確認是否有輸入資料"
    else:
        form = forms.Optiontest()
    return render(request, 'imputation.html', locals())
#-------------------------------------------------------------------------------------

@login_required(login_url='/login')
def reference_page(request):
    if request.user.is_authenticated:
        username = request.user.username

    if request.POST:
        form = forms.Makeref(request.POST, request.FILES)
        if form.is_valid():
             message = "已收到資料"
             #表單資料
             project_name = form.cleaned_data['user_name']
             path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
             num = 2
             while os.path.isdir(path2):
                 project_name = project_name+"_"+str(num)
                 path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
                 num+=1
                 if num > 50 :
                     print('error')
                     break
             handle_uploaded_bedfile(request.FILES['bedfile'],username,project_name)
             handle_uploaded_bimfile(request.FILES['bimfile'],username,project_name)
             handle_uploaded_famfile(request.FILES['famfile'],username,project_name)
             user_chromosome = form.cleaned_data['user_chromosome']
             
             user_dir = "../../../work1782/evan/reference/"+username
             
             process = "converting"

             with open(path2+"/process.txt", "w") as f:
                 f.write(process)
 

             
             creat_vcf = subprocess.Popen("./software/plink_dir/plink --allow-extra-chr --allow-no-sex --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload --recode vcf --keep-allele-order --out ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload", shell=True)
             creat_vcf.communicate()
             
             if not os.path.isdir(user_dir):
                 os.mkdir(user_dir)
             creat_reference = subprocess.Popen("perl ./code/vcf2impute_legend_haps -vcf ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload.vcf -leghap ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/user_ref -chr "+user_chromosome+" -snps-only",shell=True)
             creat_reference.communicate()
             
             #os.system("./software/zip_dir/bin/zip -r ./sample_dir/"+username+"/"+project_name+"/user_ref.zip ./sample_dir/"+username+"/"+project_name+"/user_ref.legend")
             process = "Done"

             with open(path2+"/process.txt", "w") as f:
                 f.write(process)
             
              
             return HttpResponseRedirect('/')
             

            
        else:
             message = "請確認是否有輸入資料"
    else:
        form = forms.Makeref()
    return render(request, 'reference.html', locals())
    
#split--------------------------------------------------------------------------
@login_required(login_url='/login')
def split_chr(request):
    if request.user.is_authenticated:
        username = request.user.username

    if request.POST:
        form = forms.Split(request.POST, request.FILES)
        if form.is_valid():
             message = "已收到資料"
             #表單資料
             project_name = form.cleaned_data['user_name']
             path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
             num = 2
             while os.path.isdir(path2):
                 project_name = project_name+"_"+str(num)
                 path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
                 num+=1
                 if num > 50 :
                     print('error')
                     break
             data_name = form.cleaned_data['data_name']
             handle_uploaded_bedfile(request.FILES['bedfile'],username,project_name)
             handle_uploaded_bimfile(request.FILES['bimfile'],username,project_name)
             handle_uploaded_famfile(request.FILES['famfile'],username,project_name)
             
             user_dir = "../../../work1782/evan/split_chr/"+username             
             user_dir2 = "../../../work1782/evan/split_chr/"+username+"/"+project_name
             

             if not os.path.isdir(user_dir):
                 os.mkdir(user_dir)
                 
             if not os.path.isdir(user_dir2):
                 os.mkdir(user_dir2)
             
             process = "spliting"

             with open(user_dir2+"/process.txt", "w") as f:
                 f.write(process)             
             for chr in range(1,23):
                 split_chr = subprocess.Popen("./software/plink_dir/plink --allow-extra-chr --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload --chr "+str(chr)+" --make-bed --out "+user_dir2+"/"+data_name+"_chr"+str(chr),shell=True)
                 split_chr.communicate()
             os.system("./software/zip_dir/bin/zip -r "+user_dir2+"/"+project_name+".zip "+user_dir2+"/*")
             
             process = "Done"
             
             with open(user_dir2+"/process.txt", "w") as f:
                 f.write(process)  

             
              
             return HttpResponseRedirect('/')
             

            
        else:
             message = "請確認是否有輸入資料"
    else:
        form = forms.Split()
    return render(request, 'split_chr.html', locals())    
    
#lift-------------------------------------------------------------------
@login_required(login_url='/login')
def liftover(request):
    if request.user.is_authenticated:
        username = request.user.username

    if request.POST:
        form = forms.Liftover(request.POST, request.FILES)
        if form.is_valid():
             message = "已收到資料"
             #表單資料
             project_name = form.cleaned_data['user_name']
             path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
             num = 2
             while os.path.isdir(path2):
                 project_name = project_name+"_"+str(num)
                 path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
                 num+=1
                 if num > 50 :
                     print('error')
                     break
             data_name = form.cleaned_data['data_name']
             handle_uploaded_bedfile(request.FILES['bedfile'],username,project_name)
             handle_uploaded_bimfile(request.FILES['bimfile'],username,project_name)
             handle_uploaded_famfile(request.FILES['famfile'],username,project_name)
             function = form.cleaned_data['function']
             user_dir = "../../../work1782/evan/liftover/"+username             
             user_dir2 = "../../../work1782/evan/liftover/"+username+"/"+project_name
             if not os.path.isdir(user_dir):
                 os.mkdir(user_dir)
                 
             if not os.path.isdir(user_dir2):
                 os.mkdir(user_dir2)          

             process = "Converting"
             
             
             convert_to_ped = subprocess.Popen("./software/plink_dir/plink --bfile "+path2+"/upload --recode --tab --out "+path2+"/upload_2",shell=True)
             with open(user_dir2+"/process.txt", "w") as f:
                 f.write(process)
             if function == "To 19":
                 liftover = subprocess.Popen("./software/liftover_dir/liftOverPlink.py -m "+path2+"/upload_2.map -p "+path2+"/upload_2.ped -c ./software/liftover_dir/hg38ToHg19.over.chain.gz -e ./software/liftover_dir/liftOver -o "+user_dir2+"/after_lift",shell=True)
                 liftover.communicate()
             elif function == "To 38":
                 liftover = subprocess.Popen("./software/liftover_dir/liftOverPlink.py -m "+path2+"/upload_2.map -p "+path2+"/upload_2.ped -c ./software/liftover_dir/hg19ToHg38.over.chain.gz -e ./software/liftover_dir/liftOver -o "+user_dir2+"/"+dataname+"",shell=True)
                 liftover.communicate()

             process = "Done"
             
             with open(user_dir2+"/process.txt", "w") as f:
                 f.write(process)                    

             
              
             return HttpResponseRedirect('/')
             

            
        else:
             message = "請確認是否有輸入資料"
    else:
        form = forms.Liftover()
    return render(request, 'liftover.html', locals())    

@login_required(login_url='/login')
def qcoption_all(request):
    if request.user.is_authenticated:
        username = request.user.username

    if request.method == 'POST':
        form = forms.Optiontest(request.POST, request.FILES)
        if form.is_valid():
             message = "已收到資料"
             user_name = form.cleaned_data['user_name']
             handle_uploaded_bedfile(request.FILES['bedfile'])
             handle_uploaded_bimfile(request.FILES['bimfile'])
             handle_uploaded_famfile(request.FILES['famfile'])
             user_mafoption = form.cleaned_data['user_mafoption']
             user_genooption = form.cleaned_data['user_genooption']
             user_mindoption = form.cleaned_data['user_mindoption']
             user_relatedness = form.cleaned_data['user_relatedness']
             user_sexcheck = form.cleaned_data['user_sexcheck']
             user_group = form.cleaned_data['user_group']

             if user_relatedness == True:
                 R = ' --genome --min 0.2'
             else:
                 R = ''
             if user_sexcheck == True:
                 S = ' --check-sex'
             else:
                 S = ''
             if user_mafoption == '不使用':
                 user_mafoption = ''
             else:
                 user_mafoption = user_mafoption
             plink = "./software/plink_dir/plink --bfile ./sample_dir/test"+user_mafoption+user_genooption+user_mindoption+R+S+ " --make-bed --out ./result_data/all_qc"
             os.system(plink)
             word = "Quality control OK"
             shapeit_1 ="./software/shapeit_dir/bin/shapeit -check -B ./result_data/all_qc -M ./ref_dir/genetic_map_chr22_combined_b37.txt --input-ref ./ref_dir/1000GP_Phase3_chr22.hap.gz ./ref_dir/1000GP_Phase3_chr22.legend.gz ./ref_dir/1000GP_Phase3.sample --output-log ./result_data/chr22"
             os.system(shapeit_1)
             shapeit_2 ="./software/shapeit_dir/bin/shapeit -B ./result_data/all_qc -M ./ref_dir/genetic_map_chr22_combined_b37.txt --input-ref ref_dir/1000GP_Phase3_chr22.hap.gz ref_dir/1000GP_Phase3_chr22.legend.gz ref_dir/1000GP_Phase3.sample --exclude-snp ./result_data/chr22.snp.strand.exclude --include-grp group.list -O ./result_data/chr_phased"
             os.system("echo "+user_group+" > group.list")
             os.system(shapeit_2)
             word = word+" Pre-phasing OK "
             impute = "./software/impute2_dir/impute2 -use_prephased_g -known_haps_g ./result_data/chr_phased.haps -h ./ref_dir/1000GP_Phase3_chr22.hap.gz -l ./ref_dir/1000GP_Phase3_chr22.legend.gz -m ./ref_dir/genetic_map_chr22_combined_b37.txt -int 20.4e6 20.5e6 -Ne 20000 -o ./result_data/imputed_data"
             os.system(impute)
             gtool = "./software/gtool_dir/gtool -G --g ./result_data/imputed_data --s ./result_data/chr_phased.sample --ped ./result_data/out.ped --map ./result_data/out.map --phenotype phenotype_1 --threshold 0.9"
             os.system(gtool)
             print("OK")
             return HttpResponseRedirect('/success_page/')

        else:
             message = "請確認是否有輸入資料"
    else:
        form = forms.Optiontest()
    return render(request, 'qc_all.html', locals())


@login_required(login_url='/login')
def result_page(request):
    if request.user.is_authenticated:
        username = request.user.username
    user_dir = "../../../work1782/evan/result_data/"+username
    
    if not os.path.isdir(user_dir):
        os.mkdir(user_dir)
    result = os.listdir(user_dir)
    if request.is_ajax():
        if request.method == 'POST':
            select_project = request.POST.get('data')
            try:
                if select_project: request.session['select_project'] = select_project
            except:
                pass
    
    return render(request, 'result_page.html', locals())
    
@login_required(login_url='/login')
def result_page_v2(request):
    if request.user.is_authenticated:
        username = request.user.username
    messages.get_messages(request)
    user_dir = "../../../work1782/evan/result_data/"+username
    user_dir2 = "../../../work1782/evan/reference/"+username
    user_dir3 = "../../../work1782/evan/split_chr/"+username
    user_dir4 = "../../../work1782/evan/liftover/"+username
    
    if not os.path.isdir(user_dir):
        os.mkdir(user_dir)
    result = os.listdir(user_dir)
    
    if not os.path.isdir(user_dir2):
        os.mkdir(user_dir2)
    result2 = os.listdir(user_dir2)
    
    if not os.path.isdir(user_dir3):
        os.mkdir(user_dir3)
    result3 = os.listdir(user_dir3)
    
    if not os.path.isdir(user_dir4):
        os.mkdir(user_dir4)
    result4 = os.listdir(user_dir4)
    
    if request.method =='POST':
        if 'imputation' in request.POST:
            message = "Get data"
            user_select = request.POST['imp_selection']
            print(message)
            print(user_select)
            if user_select == "":
                messages.add_message(request, messages.WARNING,'Please check the selection')
                return HttpResponseRedirect('/result_page_v2')
            else:
                request.session['select_project'] = user_select
                request.session['function'] = 'imputation'
                return HttpResponseRedirect('/Download_page_v2')
            
        if 'reference' in request.POST:
            message = "Get data"
            user_select = request.POST['ref_selection']
            print(message)
            print(user_select)                          
            if user_select == "":
                messages.add_message(request, messages.WARNING,'Please check the selection')
                return HttpResponseRedirect('/result_page_v2')
            else:
                request.session['select_project'] = user_select
                request.session['function'] = 'reference'
                return HttpResponseRedirect('/Download_page_v2')
                
        if 'split_chr' in request.POST:
            message = "Get data"
            user_select = request.POST['split_selection']
            print(message)
            print(user_select)                          
            if user_select == "":
                messages.add_message(request, messages.WARNING,'Please check the selection')
                return HttpResponseRedirect('/result_page_v2')
            else:
                request.session['select_project'] = user_select
                request.session['function'] = 'split_chr'
                return HttpResponseRedirect('/Download_page_v2')
                
        if 'liftover' in request.POST:
            message = "Get data"
            user_select = request.POST['liftover_selection']
            print(message)
            print(user_select)                          
            if user_select == "":
                messages.add_message(request, messages.WARNING,'Please check the selection')
                return HttpResponseRedirect('/result_page_v2')
            else:
                request.session['select_project'] = user_select
                request.session['function'] = 'liftover'
                return HttpResponseRedirect('/Download_page_v2')
             
    else:
        print("no data")
        message = "Please check the selection"
    
    return render(request, 'result_page_v2.html', locals())
    
    
#--------------------------------------------------------------------------------

@login_required(login_url='/login')
def download_page(request):
    if request.user.is_authenticated:
        username = request.user.username
    user_dir = "../../../work1782/evan/result_data/"+username
    
    if 'select_project' in request.session:
        select_project = request.session['select_project']
        user_dir2 = "../../../work1782/evan/result_data/"+username+"/"+select_project
        zip_dir = user_dir2+"/"+select_project+".zip"
        process_file = f"{user_dir2}/process.txt"
        if os.path.isfile(process_file):
            with open(user_dir2+"/process.txt", "r") as f:
                process = f.readline()
        info_plot = f"/home/evan/ch08www/mysite/static/user_plot/{username}/{select_project}/info.png"
        if os.path.isfile(info_plot):
            picture = f"user_plot/{username}/{select_project}/info.png"
        if os.path.isfile(zip_dir):
            zip_file = zip_dir
    if 'function' in request.session:
        function = request.session['function']
        if function == 'imputation':
            option_list = []
            with open(user_dir2+'/option.txt', 'r') as f:
                for line in f:
                    #print(line)
                    x = line.replace('\n','')
                    if x == '' or x =='No use':
                        x = 'None'
                    option_list.append(x)
            title_list = ['Date','Project name','Chromosome number','Impute the full chromosome','Start position','End position', 'Minor allele frquency option','Single SNP missing rate option','Individual SNP missing rate option','Hardy-Weinberg Equilibrium option','Reference group select']
            d = {}
            for t,c in zip(title_list,option_list):
                d[t] = c
            Date = d['Date']
            Project_name = d['Project name']
            Chr_num = d['Chromosome number']
            Entire_chr = d['Impute the full chromosome']
            Sta_posi = d['Start position']
            End_posi = d['End position']
            Maf = d['Minor allele frquency option']
            Sin_miss = d['Single SNP missing rate option']
            Indi_miss = d['Individual SNP missing rate option']
            Hardy = d['Hardy-Weinberg Equilibrium option']
            Ref = d['Reference group select']

    return render(request, 'Download_page.html', locals())

#--------------------------------------------------------------------------------檔案處理
def impute5_pipeline():
    impute5 = subprocess.Popen("./software/impute5/impute_v5.1/impute5 --h ../../../work1782/evan/1000G_vcf/impute5_ref/ALL.chr"+user_chr+".phase3_shapeit2_mvncall_integrated_v5a.20130502.genotypes.imp5 --m ../../../work1782/evan/1000G_vcf/impute5_genetic_map/chr"+user_chr+".b37.gmap.gz --g sample --r "+user_chr+" --o "+data_path+"/ouput.vcf.gz --l "+data_path+"/log/impute5.log")

def handle_uploaded_bedfile(f,username,project_name):
    path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
    if not os.path.isdir(path2):
        os.mkdir(path2)
    with open(path2+'/upload.bed', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
def handle_uploaded_bimfile(f,username,project_name):
    path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
    if not os.path.isdir(path2):
        os.mkdir(path2)
    with open(path2+'/upload.bim', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
def handle_uploaded_famfile(f,username,project_name):
    path2 = "../../../work1782/evan/sample_dir/"+username+"/"+project_name
    if not os.path.isdir(path2):
        os.mkdir(path2)
    with open(path2+'/upload.fam', 'wb+') as destination:
        for chunk in f.chunks():
            destination.write(chunk)
            
def submit_page(request):
    if request.user.is_authenticated:
        username = request.user.username
    return render(request, 'success.html', locals())
    
            
def success_page(request):
    if request.user.is_authenticated:
        username = request.user.username
    
#取出session資料    
    if 'project_name' in request.session:
        project_name = request.session['project_name']
    if 'user_chromosome' in request.session:
        user_chromosome = request.session['user_chromosome']
    if 'user_entirechr' in request.session:
        user_entirechr = request.session['user_entirechr']
    if 'user_position1' in request.session:
        user_position1 = request.session['user_position1']
    if 'user_position2' in request.session:
        user_position2 = request.session['user_position2']
    if 'user_mafoption' in request.session:
        user_mafoption = request.session['user_mafoption']
    if 'user_genooption' in request.session:
        user_genooption = request.session['user_genooption']
    if 'user_mindoption' in request.session:
        user_mindoption = request.session['user_mindoption']
    if 'user_hwe' in request.session:
        user_hwe = request.session['user_hwe']
    if 'user_group' in request.session:
        user_group = request.session['user_group']
        
#------------------------------------------------------------------------        
    #創建資料夾及pathway
    path = "../../../work1782/evan/result_data/"+username
    if not os.path.isdir(path):
        os.mkdir(path)
    data_path = "../../../work1782/evan/result_data/"+username+"/"+project_name
    os.mkdir(data_path)
    os.mkdir(data_path+"/log")
    os.mkdir(data_path+"/zip_file")
    date1 = datetime.datetime.now()
    date1 = date1.strftime("%Y-%m-%d %H:%M")
    x = print(date1)
    plot_dir = "/home/evan/ch08www/mysite/static/user_plot/"+username
    if not os.path.isdir(plot_dir):
        os.mkdir(plot_dir)    
    plot_dir2 = plot_dir+"/"+project_name
    if not os.path.isdir(plot_dir2):
        os.mkdir(plot_dir2)
    #儲存全部選項
    option_list = [x,project_name,user_chromosome,user_entirechr,user_position1,user_position2,user_mafoption,user_genooption,user_mindoption,user_hwe,user_group]
    filename = data_path+"/option.txt"
    with open(filename,'w') as f:
        for i in option_list:
            f.write(str(i)+'\n')
    #save_option = open(data_path+"/option.txt",'w')
    #print(option_list,file = save_option)
    #save_option.close()  
    
    #if user_relatedness == True:
    #    R = ' --genome --min 0.2'
    #else:
    #    R = ''
    #if user_sexcheck == True:
    #    S = ' --check-sex'
    #else:
    #    S = ''
    if user_mafoption == 'No use':
        user_mafoption = ''
    else:
        user_mafoption = user_mafoption
    if user_genooption == 'No use':
        user_genooption = ''
    else:
        user_genooption = user_genooption
    if user_mindoption == 'No use':
        user_mindoption = ''
    else:
        user_mindoption = user_mindoption
        
    if user_hwe == 'No use':
        user_hwe = ''
    else:
        user_hwe = user_hwe
        
    if user_group != 'TWB':
        ref_map = "../../../work1782/evan/ref_dir/genetic_map_chr"+user_chromosome+"_combined_b37.txt"
        ref_hap = "../../../work1782/evan/ref_dir/1000GP_Phase3_chr"+user_chromosome+".hap.gz"
        ref_legend = "../../../work1782/evan/ref_dir/1000GP_Phase3_chr"+user_chromosome+".legend.gz"
        ref_sample = "../../../work1782/evan/ref_dir/1000GP_Phase3.sample"
        include = "--include-grp "+data_path+"/group.list "
    elif user_group == 'TWB':
        ref_map = "../../../work1782/evan/ref_dir/TWB_data/chr"+user_chromosome+"_map.txt"
        ref_hap = "../../../work1782/evan/ref_dir/TWB_data/chr"+user_chromosome+"_new.hap.gz"
        ref_legend = "../../../work1782/evan/ref_dir/TWB_data/chr"+user_chromosome+"_new.legend.gz"
        ref_sample = "../../../work1782/evan/ref_dir/TWB_data/TWB_ngs.sample"
        include = ""
                    
#壓縮區----------------------------------------------------------------------------------
    compress = "./software/zip_dir/bin/zip -jr "+data_path+"/"+project_name+".zip "+data_path+"/zip_file/*"
#PLINK-----------------------------------------------------------------------------------
    try:
        process = "Quality Control"
            
        with open(data_path+"/process.txt", "w") as f:
            f.write(process)
        plink = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload"+user_mafoption+user_genooption+user_mindoption+user_hwe+" --make-bed --out "+data_path+"/all_qc", shell=True)
        plink.communicate()
        word = "Quality control OK"
    except:
        print("plink error")
#SHAPEIT1--------------------------------------------------------------------------------
    try:
        process = "Pre-phasing(1/3)"
            
        with open(data_path+"/process.txt", "w") as f:
            f.write(process)
        if user_group == 'ALL':
            include = ""
        group = open(data_path+"/group.list",'w')
        print(user_group,file = group)
        group.close()
        shapeit1 = subprocess.Popen("./software/shapeit_dir/bin/shapeit -check -B "+data_path+"/all_qc "+"-M "+ref_map+" --input-ref "+ref_hap+" "+ref_legend+" "+ref_sample+" "+include+"-T 12 --output-log "+data_path+"/log/chr"+user_chromosome+"_s1",shell=True)
        shapeit1.communicate()
        fname = data_path+"/log/chr"+user_chromosome+"_s1.log"
        with open(fname, 'r') as f:  #打開文件
            lines = f.readlines() #讀取所有行
            last_line = lines[-1][0:21] #取最後一行     
            s1_mess = last_line
            print(s1_mess)
            if s1_mess == "ERROR: Duplicate site":
                move_dup = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload"+" --list-duplicate-vars ids-only suppress-first"+" --out "+data_path+"/dupvar_exclude", shell=True)
                move_dup.communicate()
                move_dup_2 = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload"+" --exclude "+data_path+"/dupvar_exclude.dupvar"+" --make-bed --snps-only just-acgt --out ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload_rmdup", shell=True)
                move_dup_2.communicate()
                plink = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload_rmdup"+user_mafoption+user_genooption+user_mindoption+R+S+" --make-bed --out "+data_path+"/all_qc", shell=True)
                plink.communicate()
                shapeit1 = subprocess.Popen("./software/shapeit_dir/bin/shapeit -check -B "+data_path+"/all_qc "+"-M "+ref_map+" --input-ref "+ref_hap+" "+ref_legend+" "+ref_sample+" "+include+"-T 12 --output-log "+data_path+"/log/chr"+user_chromosome+"_s1",shell=True)
                shapeit1.communicate()
    except:
        print("alignment error")
        try:
            fname = data_path+"/log/chr"+user_chromosome+"_s1.log"
            with open(fname, 'r') as f:  #打開文件
                lines = f.readlines() #讀取所有行
                last_line = lines[-1][0:21] #取最後一行     
                s1_mess = last_line
                print(s1_mess,"2")
                if s1_mess == "ERROR: Duplicate site":
                    move_dup = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload"+" --list-duplicate-vars ids-only suppress-first"+" --out "+data_path+"/dupvar_exclude", shell=True)
                    move_dup.communicate()
                    move_dup_2 = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload"+" --exclude "+data_path+"/dupvar_exclude.dupvar"+" --snps-only just-acgt --out ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload_rmdup", shell=True)
                    move_dup_2.communicate()
                    plink = subprocess.Popen("./software/plink_dir/plink --bfile ../../../work1782/evan/sample_dir/"+username+"/"+project_name+"/upload_rmdup"+user_mafoption+user_genooption+user_mindoption+R+S+" --make-bed --out "+data_path+"/all_qc", shell=True)
                    plink.communicate()
                    shapeit1 = subprocess.Popen("./software/shapeit_dir/bin/shapeit -check -B "+data_path+"/all_qc "+"-M "+ref_map+" --input-ref "+ref_hap+" "+ref_legend+" "+ref_sample+" "+include+"-T 12 --output-log "+data_path+"/log/chr"+user_chromosome+"_s1",shell=True)
                    shapeit1.communicate()
        except:
            print("Shapeit other error")
#SHAPEIT2---------------------------------------------------------------------------------
    try:
        process = "Pre-phasing(2/3)"
            
        with open(data_path+"/process.txt", "w") as f:
            f.write(process)
        shapeit2 = subprocess.Popen("./software/shapeit_dir/bin/shapeit -check -B "+data_path+"/all_qc "+"-M "+ref_map+" --input-ref "+ref_hap+" "+ref_legend+" "+ref_sample+" --exclude-snp "+data_path+"/log/chr"+user_chromosome+"_s1.snp.strand.exclude "+include+"-T 12 --output-log "+data_path+"/log/chr"+user_chromosome+"_s2",shell=True)
        shapeit2.communicate()
    except:
        print("strand check error")
#SHAPEIT3--------------------------------------------------------------------------------
    try:
        process = "Pre-phasing(3/3)"
            
        with open(data_path+"/process.txt", "w") as f:
            f.write(process)
        shapeit3 = subprocess.Popen("./software/shapeit_dir/bin/shapeit -B "+data_path+"/all_qc "+"-M "+ref_map+" --input-ref "+ref_hap+" "+ref_legend+" "+ref_sample+" -T 12 --exclude-snp "+data_path+"/log/chr"+user_chromosome+"_s1.snp.strand.exclude "+include+"--output-log "+data_path+"/log/chr"+user_chromosome+"_s3 -O "+data_path+"/chr"+user_chromosome+"_phased",shell=True)
        shapeit3.communicate()
        word = word+" Pre-phasing OK "
    except:
        print("pre-phasing error")
#IMPUTE2---------------------------------------------------------------------------------
    process = "Imputation"
            
    with open(data_path+"/process.txt", "w") as f:
        f.write(process)
    if user_entirechr == True:
        with open (ref_map, 'r') as fp:
            lines = fp.readlines()
            last_line = lines[-1]
            maxPos = int(last_line.split()[0])
            numchunk = int(maxPos/5000000)
            numchunk2 = numchunk+1
            fp.close
        start = 0
        maxcpu = 10
        semlock = threading.BoundedSemaphore(maxcpu)
        threads = []
        for chunk in range(1,numchunk2+1):
            endpo = start + 5000000
            startpo = start + 1
            semlock.acquire()
            threads.append(threading.Thread(target = impute2, args = (data_path,user_chromosome,ref_hap,ref_legend,ref_map,startpo,endpo,chunk,semlock,)))
            threads[chunk-1].start()
            start = endpo
        for chunk in range(1,numchunk2+1):
            threads[chunk-1].join()
        print("impute2 done")
    else:
        maxPos = int(user_position2) - int(user_position1)
        numchunk = int(maxPos/5000000)
        numchunk2 = numchunk+1
        if user_position1 == 0:
            start = int(user_position1)
        else:
            start = int(user_position1) - 1
        for chunk in range(1,numchunk2+1):
            endpo = start + 5000000
            if endpo > int(user_position2):
                endpo = int(user_position2)
            startpo = start + 1
            impute = subprocess.Popen("./software/impute2_dir/impute2 -use_prephased_g -known_haps_g "+data_path+"/chr"+user_chromosome+"_phased.haps -h "+ref_hap+" -l "+ref_legend+" -m "+ref_map+" -int "+str(startpo)+" "+str(endpo)+" -Ne 20000 -o "+data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)+"",shell=True)
            impute.communicate()
            start = endpo
#GTOOL+MERGE-----------------------------------------------------------------------------
    exidir = ""
    process = "Converting format"
            
    with open(data_path+"/process.txt", "w") as f:
        f.write(process)
    
    for chunk in range(1,numchunk2+1):
        chunk_dir =  data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)
        if os.path.isfile(chunk_dir):
            exidir = exidir+" "+chunk_dir+".ped "+chunk_dir+".map"+"\n"
        else:
            continue
        gtool = subprocess.Popen("./software/gtool_dir/gtool -G --g "+data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)+" --s  "+data_path+"/chr"+user_chromosome+"_phased.sample --ped "+data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)+".ped --map "+data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)+".map --phenotype phenotype_1 --threshold 0.9",shell=True)
        gtool.communicate()  
    chunklist = open(""+data_path+"/list.txt",'w')
    print(exidir,file = chunklist)
    chunklist.close()
    for chunk in range(1,numchunk2+1):
        chunk_dir = data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)
        if os.path.isfile(chunk_dir):
            break
        else:
            continue
    print(chunk)
    merge = subprocess.Popen("./software/plink_dir/plink --noweb --file "+chunk_dir+" --merge-list "+data_path+"/list.txt --out "+data_path+"/zip_file/imputed_chr"+user_chromosome+"",shell=True)
    merge.communicate()
    os.system(compress)
    print("OK")
    date2 = datetime.datetime.now()
    print(date1)
    print(date2)
    process = "Done"
            
    with open(data_path+"/process.txt", "w") as f:
        f.write(process)
#轉換excel        
    for i in range(1,51): # chunk 1~50 
        inputfileTxt = data_path+"/imputed_chr"+user_chromosome+".chunk"+str(i)+"_info"  # path and all chunks filename, can just change chr1 to chr2 ...
        outfileExcel = data_path+"/chunk"+str(i)+".xlsx" #output_path
        if os.path.isfile(inputfileTxt):
            txt_to_xlsx(inputfileTxt,outfileExcel)
            print("chunk",i,"done")
    print("-------------")
    print("Conver excel DONE")
    
    #os.system(f"cp /home/evan/ch08www/code/info_plot.py {data_path}/")
    #print(f"{data_path}")
    os.system(f"python /home/evan/ch08www/code/info_plot.py {data_path} {plot_dir2}")
    #os.system(f"cp {data_path}/info.png {plot_dir2}")    
    print("PLOT DONE")
    
    return HttpResponseRedirect('/')

    return render(request, 'index.html', locals())
#------------------------------------------------------------------------------檔案下載
@login_required(login_url='/login')
def file_download(request):
    if request.user.is_authenticated:
        username = request.user.username
    if 'select_project' in request.session:
        select_project = request.session['select_project']
    if 'function' in request.session:
        function = request.session['function']
    
    if function == 'imputation':
        download_dir = '../../../work1782/evan/result_data/'
    elif function == 'reference':
        download_dir = '../../../work1782/evan/reference/'
    elif function == 'split_chr':
        download_dir = '../../../work1782/evan/split_chr/'
    elif function == 'liftover':
        download_dir = '../../../work1782/evan/liftover/'
    print(select_project)
    filename= download_dir+str(username)+"/"+str(select_project)+"/"+str(select_project)+".zip"
    file = open((filename),'rb')
    response = FileResponse(file)
    response['Content-Type']='application/octet-stream'
    response['Content-Disposition']="attachment;filename="+select_project+".zip"
    return response    

def impute2(data_path,user_chromosome,ref_hap,ref_legend,ref_map,startpo,endpo,chunk,semlock):
    impute = subprocess.Popen("./software/impute2_dir/impute2 -use_prephased_g -known_haps_g "+data_path+"/chr"+user_chromosome+"_phased.haps -strand_g "+data_path+"/log/chr"+user_chromosome+"_s1.snp.strand -h "+ref_hap+" -l "+ref_legend+" -m "+ref_map+" -int "+str(startpo)+" "+str(endpo)+" -Ne 20000 -o "+data_path+"/imputed_chr"+user_chromosome+".chunk"+str(chunk)+"",shell=True)
    impute.communicate()  
    semlock.release()
    
    
@login_required(login_url='/login')
def file_download_ref(request):
    if request.user.is_authenticated:
        username = request.user.username
        
    filename= "./sample_dir//"+str(username)+"/"+str(project_name)+"/"+str(project_name)+".zip"
    file = open((filename),'rb')
    response = FileResponse(file)
    response['Content-Type']='application/octet-stream'
    response['Content-Disposition']="attachment;filename="+project_name+".zip"
    return response    
    #if request.user.is_authenticated:
    #    username = request.user.username
    #    file=open(file_path+"/"+project_name+".zip",'rb')
    #    response =FileResponse(file)
    #    response['Content-Type']='application/octet-stream'
    #    response['Content-Disposition']="attachment;filename="+project_name+".zip"
    
    #return response
#    try:
#        file = open(file_path+"/"+project_name+".zip",'rb')
#        data = file.read()
#        file.close()
#        response = HttpResponse(data, content_type='application/zip')
#        response['Content-Disposition']="attachment;filename="+project_name+".zip"
#        from urllib import parse
#        response['Content-Disposition'] = 'attachment;filename='   parse.quote(z_name)
#        return response
#    except Exception as e:
#        print(e)

